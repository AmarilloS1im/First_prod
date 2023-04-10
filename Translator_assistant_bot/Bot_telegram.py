import copy
import csv
import logging
import os
import re
import shutil
import time
import uuid
import openpyxl
import emoji
from aiogram import Bot, Dispatcher, executor, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import ContentType
import pathlib


# <editor-fold desc="Dictionary class">
class TranslatedWordsDictMeta:

    def __init__(self, *args, **kwargs):
        self.__main_dict = None
        self.__custom_dict = None

    @property
    def main_dict(self):
        my_dict = []
        with open(rf"main_dict\translated_words_dict_file.csv") as r_file:
            file_reader = csv.reader(r_file, delimiter=";")
            for row in file_reader:
                tmp_dict = [row[0], row[1]]
                my_dict.append(tmp_dict)
        self.__main_dict = list(sorted(my_dict, key=lambda item: len(item[0]), reverse=True))
        return self.__main_dict

    def set_custom_dict(self, value):
        my_dict = []
        with open(value, 'r', encoding='utf-8-sig') as r_file:
            file_reader = csv.reader(r_file, delimiter=";")
            for row in file_reader:
                tmp_dict = [row[0], row[1]]
                my_dict.append(tmp_dict)
        self.__custom_dict = list(sorted(my_dict, key=lambda item: len(item[0]), reverse=True))

    def get_custom_dict(self):
        return self.__custom_dict


class WordsDict(metaclass=TranslatedWordsDictMeta):
    pass


# </editor-fold>

# <editor-fold desc="UUID class">
class UUIDict:
    def __init__(self, file_from_user):
        self.__file_from_user = file_from_user
        self.__new_key = str(uuid.uuid4())
        self.__uuid_dict = {}
        self.__uuid_dict = self.set_uuid_dict(self.get_key, self.get_file_from_user)

    @property
    def get_file_from_user(self):
        return self.__file_from_user

    @property
    def get_key(self):
        return self.__new_key

    @property
    def uuid_dict(self):
        return self.__uuid_dict

    def set_uuid_dict(self, key, value):
        if key in self.__uuid_dict.keys():
            key = str(uuid.uuid4())
        else:
            self.__uuid_dict[key] = value
            return self.__uuid_dict


# </editor-fold>


# <editor-fold desc="UserFile class">
class UserFile:
    def __init__(self, user_file):
        self.__user_file = user_file
        self.__user_file_name = None
        self.__user_file_extension = None

    @property
    def get_user_file(self):
        return self.__user_file

    @property
    def get_user_file_extension(self):
        self.__user_file_extension = pathlib.PurePosixPath(self.__user_file).suffix
        return self.__user_file_extension

    @property
    def get_user_file_name(self):
        self.__user_file_name = re.split(self.get_user_file_extension, self.get_user_file)
        return self.__user_file_name[0]


# </editor-fold>

# <editor-fold desc="FSM admin class">
class FSMAdmin(StatesGroup):
    file_to_translate = State()
    add_custom_dict = State()


# </editor-fold>


# <editor-fold desc="Translarion block">
def start_translate(file_name, user_id):
    file_name = file_name
    user_id = user_id
    make_translated_doc(get_translated_list(get_info_from_doc(file_name), user_id), file_name)


def get_info_from_doc(file_name):
    book = openpyxl.open(rf"translated_docs\{file_name}", read_only=True, data_only=True)
    sheet = book.active
    data_list = []
    for row in range(2, sheet.max_row):
        temp_data_list = []
        for cell in range(1, sheet.max_column):
            temp_data_list.append(sheet[row][cell].value)
        data_list.append(temp_data_list)
    book.close()
    return data_list


def get_translated_list(data_list, user_id):
    # translate function
    def translate(str_to_translate, user_id):
        translated_string = str_to_translate[:]
        user_id_dict_name = rf"{user_id}.csv"
        if user_id_dict_name in os.listdir('custom_dict/'):
            WordsDict.set_custom_dict(rf"custom_dict/{user_id_dict_name}")
            translated_dict = WordsDict.get_custom_dict()
        else:
            translated_dict = WordsDict.main_dict
        for i in range(len(translated_dict)):
            if translated_dict[i][0] == 'М.П.' or translated_dict[i][0] == 'г.':
                translated_string = re.sub(fr"{translated_dict[i][0]}", f"{translated_dict[i][1]}", translated_string)
            else:
                translated_string = re.sub(fr"\b{translated_dict[i][0]}\b", f"{translated_dict[i][1]}",
                                           translated_string)
        return translated_string

    # make translate
    copy_data_list = copy.deepcopy(data_list)
    translated_list = []
    temp_out_list = []
    for i in range(len(data_list)):
        for j in range(len(data_list[i])):
            if not isinstance(data_list[i][j], str):
                temp_out_list.append(data_list[i][j])
            else:
                if translate(copy_data_list[i][j], user_id) == data_list[i][j]:
                    temp_out_list.append(data_list[i][j])
                else:
                    temp_out_list.append(data_list[i][j] + ' ' + "/" + ' ' + translate(copy_data_list[i][j], user_id))
        translated_list.append(temp_out_list)
        temp_out_list = []

    # align_and_adjust

    def align_and_adjust(translated_list):
        for i in range(len(translated_list)):
            for j in range(len(translated_list[i])):
                if isinstance(translated_list[i][j], str):
                    split_word = re.split(r"\s/\s", translated_list[i][j])
                    if '\n' in split_word[0]:
                        left_result = re.split(r"\n", split_word[0])
                        right_result = re.split(r"\n", split_word[-1])
                        new_word = ''
                        for x in range(len(left_result)):
                            new_word = new_word + left_result[x] + ' ' + '/' + ' ' + right_result[x] + '\n'
                        split_word[0] = new_word[:-1]
                        split_word[-1] = new_word[:-1]
                        translated_list[i][j] = split_word[0]
        return translated_list

    translated_list = align_and_adjust(translated_list)
    return translated_list


# make_translated_doc
def make_translated_doc(translated_list, file_name):
    max_row = len(get_info_from_doc(file_name))
    shutil.copy(rf"translated_docs\{file_name}", rf"translated_docs\__Translated__{file_name}")
    work_book = openpyxl.open(rf"translated_docs\__Translated__{file_name}", read_only=False, data_only=False)
    active_sheet = work_book.active

    for row in range(2, max_row):
        for cell in range(1, active_sheet.max_column):
            if translated_list[row - 2][cell - 1] is not None:
                active_sheet[row][cell].value = translated_list[row - 2][cell - 1]
    work_book.save(rf"translated_docs\__Translated__{file_name}")
    work_book.close()


# </editor-fold>

# <editor-fold desc="TelegramBot block">
# <editor-fold desc="Bot initialisation and buttons">

# Configure logging
logging.basicConfig(level=logging.INFO)

# Initialize bot and dispatcher
bot = Bot(token=os.getenv('API_TOKEN'))

storage = MemoryStorage()

dp = Dispatcher(bot, storage=storage)

# Make buttons
translate_button = types.InlineKeyboardButton('ПЕРЕВЕСТИ', callback_data='translate_button')
add_new_words_button = types.InlineKeyboardButton('ДОБАВИТЬ НОВЫЙ СЛОВАРЬ', callback_data='add_to_dict')
help_button = types.InlineKeyboardButton('ПОМОЩЬ', callback_data='help')
markup_start_screen = types.InlineKeyboardMarkup(row_width=2)
markup_start_screen.add(translate_button, add_new_words_button, help_button)

button_back = types.InlineKeyboardButton('НАЗАД', callback_data='back')
markup_back = types.InlineKeyboardMarkup(row_width=1)
markup_back.add(button_back)

markup_dicts = types.InlineKeyboardMarkup(row_width=1)
main_dict_button = types.InlineKeyboardButton('СКАЧАТЬ БАЗОВЫЙ СЛОВАРЬ', callback_data='main_dict')
custom_dict_button = types.InlineKeyboardButton('ЗАГРУЗИТЬ ПОЛЬЗОВАТЕЛЬСКИЙ СЛОВАРЬ', callback_data='set_custom_dict')
restore_main_dict_button = types.InlineKeyboardButton('УДАЛИТЬ ПОЛЬЗОВАТЕЛЬСКИЙ СЛОВОВАРЬ',
                                                      callback_data='restore_main_dict')
markup_dicts.add(main_dict_button, custom_dict_button, restore_main_dict_button, help_button, button_back)

markup_next_doc = types.InlineKeyboardMarkup(row_width=1)
next_doc_button = types.InlineKeyboardButton('ПЕРЕВЕСТИ СЛЕДУЮЩИЙ ДОКУМЕНТ', callback_data='next_doc')
markup_next_doc.add(next_doc_button, button_back)


# </editor-fold>

# <editor-fold desc="Start screen">
# START SCREEN

@dp.message_handler(commands=['start'])
async def send_welcome(message: types.Message):
    user_id = message.from_user.id
    user_nikname = message.from_user.username
    user_full_name = message.from_user.full_name
    logging.info(f'{user_id}-----{user_full_name} {user_nikname} {time.asctime()}')
    await message.reply(
        f"Привет {user_full_name}\nЯ Translator's assistant bot\nЧтобы перевести документ нажми кнопку ПЕРЕВЕСТИ\n"
        f"Чтобы добавить новые слова в словарь нажми кнопку ДОБАВИТЬ НОВЫЙ СЛОВАРЬ\n"
        f"Для получения более подробной информации нажми кнопку ПОМОЩЬ",
        reply_markup=markup_start_screen)


# </editor-fold>

# <editor-fold desc="Translate button action">
# TRANSLATE BUTTON
@dp.callback_query_handler(text='translate_button')
async def callback_translate(callback: types.CallbackQuery):
    await FSMAdmin.file_to_translate.set()
    await callback.message.answer(f"ЧТОБЫ ЗАГРУЗИТЕ ДОКУМЕНТ, КОТОРЫЙ НУЖНО ПЕРЕВЕСТИ\n"
                                  f"НАЖМИТЕ НА {emoji.emojize(':paperclip:')}", reply_markup=markup_back)


@dp.message_handler(content_types=types.ContentType.ANY, state=FSMAdmin.file_to_translate)
async def load_on_server_file_to_translate(message: types.Message, state: FSMContext):
    if message.content_type != 'document':
        await FSMAdmin.file_to_translate.set()
        await message.answer('Я умею перводить только Excel файлы, загрузите файл с раширениме .xls или .xlsx',
                             reply_markup=markup_back)
    else:
        user_file = UserFile(message.document.file_name)
        user_unique_id = message.from_user.id
        file_name_get_from_user = rf"{user_file.get_user_file_name}{user_unique_id}{user_file.get_user_file_extension}"
        uuid_data = UUIDict(file_name_get_from_user)
        uuid_user_file = rf"{uuid_data.get_key}{user_unique_id}{user_file.get_user_file_extension}"
        if user_file.get_user_file_extension != '.xlsx' and user_file.get_user_file_extension != '.xls':
            await FSMAdmin.file_to_translate.set()
            await message.answer('Документ должен быть в формате .xls или .xlsx', reply_markup=markup_back)
        else:
            await message.document.download(destination_file=rf"translated_docs\{uuid_user_file}")
            await message.answer('БОТ НАЧАЛ ПЕРВОД ДОКУМЕНТА, ОЖИДАЙТЕ')
            start_translate(uuid_user_file, user_unique_id)
            translated_uuid_doc = rf"__Translated__{uuid_data.get_key}{user_unique_id}" \
                                  rf"{user_file.get_user_file_extension}"

            original_file_data = rf"{uuid_data.uuid_dict[uuid_data.get_key]}"
            original_file_data = UserFile(original_file_data)
            original_f_name = original_file_data.get_user_file_name
            original_f_extension = original_file_data.get_user_file_extension

            shutil.copy(rf"translated_docs\{translated_uuid_doc}",
                        rf"translated_docs\archive\__Translated__{original_f_name}{original_f_extension}")

            translated_userfriendly_file = rf"translated_docs\archive\__Translated__{original_f_name}" \
                                           rf"{original_f_extension}"

            reply_translate_doc = open(translated_userfriendly_file, 'rb')
            await message.answer('ФАЙЛ ПЕРВЕДЕН И ГОТОВ К СКАЧИВАНИЮ')
            await message.reply_document(reply_translate_doc, reply_markup=markup_next_doc)

            if uuid_user_file in os.listdir(f'translated_docs'):
                os.remove(rf"translated_docs\{uuid_user_file}")
            if translated_uuid_doc in os.listdir(f'translated_docs'):
                os.remove(rf"translated_docs\{translated_uuid_doc}")
            os.remove(translated_userfriendly_file)
            await state.finish()


@dp.callback_query_handler(text='next_doc')
async def callback_next_doc_translate(callback: types.CallbackQuery):
    await FSMAdmin.file_to_translate.set()
    await callback.message.answer(f"ЧТОБЫ ЗАГРУЗИТЕ ДОКУМЕНТ, КОТОРЫЙ НУЖНО ПЕРЕВЕСТИ\n"
                                  f"НАЖМИТЕ НА {emoji.emojize(':paperclip:')}", reply_markup=markup_back)


# </editor-fold>


# Add new words to dict section

@dp.callback_query_handler(text='add_to_dict')
async def add_to_dict(callback: types.CallbackQuery):
    await callback.message.answer(
        'Чтобы добавить пользовательский словарь, загрузите файл с новыми словами в формате .CSV\n'
        'Для корректной загрузки словаря, см. справку в разделе "ПОМОЩЬ"', reply_markup=markup_dicts)


@dp.callback_query_handler(text='set_custom_dict')
async def set_custom_dicts(callback: types.CallbackQuery):
    await callback.message.answer(f"ЧТОБЫ ЗАГРУЗИТЕ ПОЛЬЗОВАТЕЛЬСКИЙ СЛОВАРЬ,"
                                  f"НАЖМИТЕ НА {emoji.emojize(':paperclip:')}", reply_markup=markup_back)
    await FSMAdmin.add_custom_dict.set()


@dp.callback_query_handler(text='main_dict')
async def upload_main_dict_to_user(callback: types.CallbackQuery):
    reply_main_dict = open(r"main_dict\translated_words_dict_file.csv", 'rb')
    await callback.message.reply_document(reply_main_dict)
    await callback.answer('Базовый словарь получен')


@dp.callback_query_handler(text='restore_main_dict')
async def restore_dict(callback: types.CallbackQuery):
    user_id = callback.from_user.id
    extension = UserFile(rf"{user_id}.csv")
    try:
        os.remove(rf"custom_dict\{user_id}{extension.get_user_file_extension}")
        await callback.message.answer('ПОЛЬЗОВАТЕЛЬСКИЙ СЛОВАРЬ УДАЛЕН, БАЗОВЫЙ СЛОВАРЬ ВОССТАНОВЛЕН')
    except:
        await callback.message.answer('У ВАС НЕТ ДЕЙСТВУЮЩЕГО ПОЛЬЗОВАТЕЛЬСКОГО СЛОВАРЯ')


@dp.message_handler(content_types=ContentType.ANY, state=FSMAdmin.add_custom_dict)
async def load_on_server_dict(message: types.Message, state: FSMContext):
    if message.content_type != 'document':
        await FSMAdmin.add_custom_dict.set()
        await message.answer('Словарем может быть только файл в формате .csv', reply_markup=markup_back)
    else:
        user_id = message.from_user.id
        file_name_dict = UserFile(message.document.file_name)
        file_name_dict_extension = file_name_dict.get_user_file_extension
        if file_name_dict_extension != '.csv':
            await FSMAdmin.add_custom_dict.set()
            await message.answer('Файл должен иметь расширение .csv', reply_markup=markup_back)
        else:
            src = rf"custom_dict\{user_id}{file_name_dict.get_user_file_extension}"
            await message.document.download(destination_file=src)
            await message.answer('ВАШ СЛОВАРЬ УСПЕШНО ДОБАВЛЕН,МОЖЕТЕ НАЧАТЬ ПЕРВОД ДОКУМЕНТА')
            await state.finish()


# <editor-fold desc="Help button">
# Help button
@dp.callback_query_handler(text='help')
async def callback_help(callback: types.CallbackQuery):
    await callback.message.reply(
        'Чтобы добавить пользовательский словарь, загрузите файл с новыми словами в формате .CSV\n'
        'СОЗДАЙТЕ ПУСТОЙ ФАЛЙ В ФОРМАТЕ .TXT\nПОСЛЕ СОЗДАНИЯ ПЕРЕИМЕНУЙТЕ ЕГО ТАК, КАК ВАМ УДОБНО И ПОМЕНЯЙТЕ '
        'РАСШИРЕНИЕ ФАЙЛА НА .CSV\n'
        'Далее откройте файл и добавьте в колонку A "слово", в колонку B "перевод".\n'
        'Количество пар(слово ; перевод) не ограниченно. ДОБАВЛЯЙТЕ КАЖДУ НОВУЮ ПАРУ СТРОЧКОЙ НИЖЕ.\n\n'
        '\t\t\t\t\t\t\t\tВНИМАНИЕ!    ATTENTION!     ACHTUNG!    \n\n'
        'ВАШ СЛОВАРЬ ДОЛЖЕН ИМЕТЬ КОДИРОВКУ "utf-8 со спецификацией"\n\n'
        'Когда ваш словарь будет готов, загрузите его с помощью кнопки ДОБАВИТЬ '
        'НОВЫЙ СЛОВАРЬ-->ЗАГРУЗИТЬ ПОЛЬЗОВАТЕЛЬСКИЙ СЛОВАРЬ"\n'
        'Вы также можете расширить базовый словарь путем добавления в него своих слов.\n'
        'Для этого в разделе ДОБАВИТЬ НОВЫЙ СЛОВАРЬ нажмите кнопку СКАЧАТЬ БАЗОВЫЙ СЛОВАРЬ\n\n'
        'Сохраните его себе на компьютер, добавьте в него необходимые вам слова,'
        ' после этого загрузите этот файл нажав на кноку '
        'ДОБАВИТЬ НОВЫЙ СЛОВАРЬ-->ЗАГРУЗИТЬ ПОЛЬЗОВАТЕЛЬСКИЙ СЛОВАРЬ\n'
        'После этого можете приступать к перводу\n'
        'ВНИМАНИЕ!\nПри нажатии на кнопку УДАЛИТЬ ПОЛЬЗОВАТЕЛЬСКИЙ СЛОВОВАРЬ, ваш словарь будет полностью удален\n'
        'Весь дальнейший первод будет идти на основание базового словаря ',
        reply_markup=markup_back)


# </editor-fold>


# <editor-fold desc="Button Back">
# Button Back

@dp.callback_query_handler(text='back', state=[FSMAdmin.add_custom_dict, FSMAdmin.file_to_translate, None])
async def callback_back_button(callback: types.CallbackQuery, state: FSMContext):
    user_full_name = callback.from_user.full_name
    await state.finish()
    await callback.message.reply(
        f'Привет {user_full_name}\nЯ DocTranslatorBot\nЧтобы перевести документ нажми кнопку ПЕРЕВЕСТИ\n'
        f'Чтобы добавить новые слова в словарь нажми кнопку ДОБАВИТЬ НОВЫЕ СЛОВА.',
        reply_markup=markup_start_screen)


# </editor-fold>


if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)
# </editor-fold>
